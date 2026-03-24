#!/usr/bin/env python3
"""
CATF <-> F2E CCS Database Merge Script
-------------------------------------
Updates CATF-sourced columns from a fresh CATF download while preserving
all F2E enrichment columns (contacts, funding, notes, SharePoint links).

Usage:
    python catf_merge.py <fresh_catf.xlsx> <f2e_database.xlsx> [--output merged.xlsx]

The script matches projects on (Project Name + Country) as the join key.
- Existing projects: CATF columns updated, F2E columns preserved
- New CATF projects: appended with empty F2E columns
- Removed CATF projects: kept in F2E database, flagged in notes
"""

import argparse
import pandas as pd
import openpyxl
from datetime import datetime
from difflib import SequenceMatcher
import sys


def classify_onshore_offshore(df):
    """Add/update 'Onshore/Offshore' column based on coordinates using Natural Earth land polygons."""
    try:
        import geopandas as gpd
        from shapely.geometry import Point

        try:
            import geodatasets
            land = gpd.read_file(geodatasets.get_path('naturalearth.land'))
        except Exception:
            land = gpd.read_file(gpd.datasets.get_path('naturalearth_land'))

        land_union = land.geometry.unary_union

        def _classify(row):
            lat = row.get('Approx. Latitude')
            lon = row.get('Approx. Longitude')
            if pd.isna(lat) or pd.isna(lon):
                return 'Unknown'
            try:
                return 'Onshore' if land_union.contains(Point(float(lon), float(lat))) else 'Offshore'
            except Exception:
                return 'Unknown'

        return df.apply(_classify, axis=1)
    except ImportError:
        print("  WARNING: geopandas not available — 'Onshore/Offshore' column will be set to 'Unknown' for new projects.")
        print("           Install with: pip install geopandas geodatasets")
        return pd.Series(['Unknown'] * len(df), index=df.index)

# Ensure print output is flushed immediately
sys.stdout.reconfigure(line_buffering=True)

CATF_COLUMNS = [
    'Project Name', 'Entities', 'Capture or Storage Details', 'Country', 'Location',
    'Sector Classification', 'Sector Description', 'Subsector Classification', 'Subsector Description',
    'Approx. Latitude', 'Approx. Longitude',
    'Capacity (Metric Tons Per Annum)',
    'Storage Classification', 'Storage Description',
    'Year Announced', 'Year Operational', 'Status', 'Notes', 'Month Announced', 'Reference'
]

F2E_COLUMNS = [
    'capture_lead_company', 'capture_lead_contact', 'capture_lead_email',
    'storage_lead_company', 'storage_lead_contact', 'storage_lead_email',
    'transport_operator', 'storage_capacity_mt', 'funding_source', 'funding_amount_eur_m',
    'eu_funding_program', 'fid_date', 'contact_status', 'contact_date', 'contact_notes',
    'internal_priority', 'relevance_f2e', 'project_website', 'sharepoint_folder',
    'internal_notes', 'last_updated_f2e', 'updated_by', 'source'
]

COUNTRY_CODES = {
    'Belgium': 'BE', 'Bulgaria': 'BG', 'Croatia': 'HR', 'Czechia': 'CZ',
    'Denmark': 'DK', 'Finland': 'FI', 'France': 'FR', 'Germany': 'DE',
    'Greece': 'GR', 'Hungary': 'HU', 'Iceland': 'IS', 'Italy': 'IT',
    'Latvia': 'LV', 'Lithuania': 'LT', 'Netherlands': 'NL', 'Norway': 'NO',
    'Poland': 'PL', 'Romania': 'RO', 'Slovakia': 'SK', 'Spain': 'ES',
    'Sweden': 'SE', 'Switzerland': 'CH', 'UK': 'GB',
}


def _parse_single_number(val):
    """Try to parse val as a single positive number. Returns the float, or None if not possible."""
    if pd.isna(val):
        return None
    s = str(val).strip().replace(',', '').replace(' ', '')
    if not s or s.lower() in ('unknown', 'unavailable', 'n/a', '-'):
        return None
    # Reject ranges (contain a dash that isn't a negative sign at the start)
    if '-' in s.lstrip('-'):
        return None
    try:
        n = float(s)
        return n if n > 0 else None
    except ValueError:
        return None


def normalize_capacity(cap_val, vis_cap_val=None):
    """Resolve best capacity value: prefer Capacity if it's a single number,
    fall back to Visualized Capacity, otherwise return 'Unknown'."""
    n = _parse_single_number(cap_val)
    if n is not None:
        return cap_val  # keep original formatting
    if vis_cap_val is not None:
        n = _parse_single_number(vis_cap_val)
        if n is not None:
            return vis_cap_val
    return 'Unknown'


FUZZY_THRESHOLD = 0.8


def make_key(row):
    return f"{str(row.get('Project Name', '')).strip().lower()}|{str(row.get('Country', '')).strip().lower()}"


def fuzzy_match_projects(catf_unmatched, f2e_unmatched):
    """Find fuzzy matches between unmatched CATF and F2E projects (same country, similar name).
    Returns list of (catf_key, f2e_key, catf_name, f2e_name, score) tuples."""
    matches = []
    # Group by country for efficiency
    catf_by_country = {}
    for key in catf_unmatched:
        name, country = key.rsplit('|', 1)
        catf_by_country.setdefault(country, []).append((key, name))

    f2e_by_country = {}
    for key in f2e_unmatched:
        name, country = key.rsplit('|', 1)
        f2e_by_country.setdefault(country, []).append((key, name))

    for country in catf_by_country:
        if country not in f2e_by_country:
            continue
        for catf_key, catf_name in catf_by_country[country]:
            best_score = 0
            best_f2e = None
            for f2e_key, f2e_name in f2e_by_country[country]:
                score = SequenceMatcher(None, catf_name, f2e_name).ratio()
                if score > best_score:
                    best_score = score
                    best_f2e = (f2e_key, f2e_name)
            if best_score >= FUZZY_THRESHOLD and best_f2e:
                matches.append((catf_key, best_f2e[0], catf_name, best_f2e[1], best_score))
    return matches


def extract_hyperlinks(xlsx_path, sheet_name, columns):
    """Return a dict of {col_name: {row_index: url}} for cells with hyperlinks."""
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb[sheet_name]
    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col_indices = {col: idx for idx, col in enumerate(header) if col in columns}
    result = {col: {} for col in col_indices}
    for row in ws.iter_rows(min_row=2):
        pandas_idx = row[0].row - 2  # 0-based pandas index
        for col, col_idx in col_indices.items():
            cell = row[col_idx]
            if cell.hyperlink and cell.hyperlink.target:
                result[col][pandas_idx] = cell.hyperlink.target
    return result


def merge(catf_path, f2e_path, output_path):
    print(f"Loading fresh CATF data from: {catf_path}")
    catf = pd.read_excel(catf_path, sheet_name='Europe')
    catf = catf.loc[:, ~catf.columns.str.startswith('Unnamed')]
    hyperlinks = extract_hyperlinks(catf_path, 'Europe', ['Reference'])
    for col, links in hyperlinks.items():
        for row_idx, url in links.items():
            catf.at[row_idx, col] = url
    catf['Status'] = catf['Status'].replace('In development', 'In Development')
    # Fix common CATF typos and case inconsistencies
    storage_fixes = {
        'Uknown': 'Unknown',
        'Depleted gas reservoir': 'Depleted Gas Reservoir',
        'Depleted onshore reservoir': 'Depleted Onshore Reservoir',
        'Depleted Oil or Gas reservoir': 'Depleted Oil or Gas Reservoir',
    }
    for col in ['Storage Classification', 'Storage Description']:
        if col in catf.columns:
            for old, new in storage_fixes.items():
                catf[col] = catf[col].str.replace(old, new, regex=False)
    if 'Capacity (Metric Tons Per Annum)' in catf.columns:
        vis_col = 'Visualized Capacity (Metric Tons Per Annum)'
        vis = catf[vis_col] if vis_col in catf.columns else pd.Series([None] * len(catf))
        catf['Capacity (Metric Tons Per Annum)'] = [
            normalize_capacity(c, v) for c, v in zip(catf['Capacity (Metric Tons Per Annum)'], vis)
        ]
    catf['_key'] = catf.apply(make_key, axis=1)
    print(f"  -> {len(catf)} projects in fresh CATF file")

    print(f"Loading existing F2E database from: {f2e_path}")
    f2e_wb_sheets = openpyxl.load_workbook(f2e_path, read_only=True).sheetnames
    f2e_sheet = 'Projects' if 'Projects' in f2e_wb_sheets else f2e_wb_sheets[0]
    print(f"  Using sheet: {f2e_sheet}")
    f2e = pd.read_excel(f2e_path, sheet_name=f2e_sheet)
    f2e['_key'] = f2e.apply(make_key, axis=1)
    print(f"  -> {len(f2e)} projects in existing F2E database")

    existing_keys = set(f2e['_key'])
    catf_keys = set(catf['_key'])
    exact_matches = existing_keys & catf_keys
    catf_unmatched = catf_keys - existing_keys
    f2e_unmatched = existing_keys - catf_keys

    # Fuzzy matching on unmatched projects
    fuzzy_matches = fuzzy_match_projects(catf_unmatched, f2e_unmatched)
    fuzzy_catf_to_f2e = {}  # catf_key -> f2e_key
    if fuzzy_matches:
        print(f"\n  Fuzzy matches (auto-accepted, threshold {FUZZY_THRESHOLD}):")
        for catf_key, f2e_key, catf_name, f2e_name, score in fuzzy_matches:
            print(f"    {score:.0%}  CATF \"{catf_name}\" <- -> F2E \"{f2e_name}\"")
            fuzzy_catf_to_f2e[catf_key] = f2e_key
            catf_unmatched.discard(catf_key)
            f2e_unmatched.discard(f2e_key)

    updated_keys = exact_matches | set(fuzzy_catf_to_f2e.values())
    new_keys = catf_unmatched
    removed_keys = f2e_unmatched

    print(f"\nMerge summary:")
    print(f"  Exact matches (CATF columns refreshed): {len(exact_matches)}")
    print(f"  Fuzzy matches (auto-accepted):          {len(fuzzy_catf_to_f2e)}")
    print(f"  New projects to add:                    {len(new_keys)}")
    print(f"  In F2E but not in CATF anymore:         {len(removed_keys)}")

    # Build CATF lookup — map both exact and fuzzy keys to CATF data
    catf_lookup = catf.set_index('_key')
    # For fuzzy matches, create a reverse lookup: f2e_key -> catf_key
    f2e_to_catf_key = {v: k for k, v in fuzzy_catf_to_f2e.items()}

    # Update existing rows with fresh CATF data
    for idx, row in f2e.iterrows():
        key = row['_key']
        lookup_key = f2e_to_catf_key.get(key, key)  # use fuzzy-mapped CATF key if applicable
        if lookup_key in catf_lookup.index:
            fresh = catf_lookup.loc[lookup_key]
            if isinstance(fresh, pd.DataFrame):
                fresh = fresh.iloc[0]
            for col in CATF_COLUMNS:
                if col in fresh.index:
                    f2e.at[idx, col] = fresh[col]
            # Set source to CATF for matched projects (if not already set)
            if pd.isna(row.get('source')) or str(row.get('source', '')).strip() == '':
                f2e.at[idx, 'source'] = 'CATF'

    # Flag removed projects — only if source is CATF (not manually added)
    today = datetime.now().strftime('%Y-%m-%d')
    for idx, row in f2e.iterrows():
        if row['_key'] in removed_keys:
            source = str(row.get('source', '') or '').strip()
            if source != 'Manual':
                existing_notes = str(row.get('internal_notes', '') or '')
                flag = f"[{today}] No longer in CATF database"
                if flag not in existing_notes:
                    f2e.at[idx, 'internal_notes'] = f"{existing_notes} {flag}".strip()

    # Add new projects
    if new_keys:
        max_counters = {}
        for pid in f2e['project_id'].dropna():
            parts = str(pid).split('-')
            if len(parts) == 2:
                cc, num = parts
                max_counters[cc] = max(max_counters.get(cc, 0), int(num))

        new_rows = []
        for _, row in catf.iterrows():
            if row['_key'] in new_keys:
                cc = COUNTRY_CODES.get(row['Country'], 'XX')
                max_counters[cc] = max_counters.get(cc, 0) + 1
                new_row = {'project_id': f"{cc}-{max_counters[cc]:03d}"}
                for col in CATF_COLUMNS:
                    new_row[col] = row.get(col, '')
                for col in F2E_COLUMNS:
                    new_row[col] = ''
                new_row['Onshore/Offshore'] = ''
                new_row['source'] = 'CATF'
                new_rows.append(new_row)

        new_df = pd.DataFrame(new_rows)
        print("  Classifying new projects as Onshore/Offshore...")
        new_df['Onshore/Offshore'] = classify_onshore_offshore(new_df)
        loc_counts = new_df['Onshore/Offshore'].value_counts().to_dict()
        print(f"    -> {loc_counts}")
        f2e = pd.concat([f2e, new_df], ignore_index=True)
        print(f"\n  Added {len(new_rows)} new projects")

    f2e = f2e.drop(columns=['_key'], errors='ignore')

    # Save by updating the existing workbook in-place to preserve formatting
    import shutil
    if output_path != f2e_path:
        shutil.copy2(f2e_path, output_path)
    wb = openpyxl.load_workbook(output_path)
    ws = wb[f2e_sheet]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col_map = {name: idx + 1 for idx, name in enumerate(headers)}  # 1-based

    # Add any new columns not yet in the sheet (e.g. 'source')
    for col_name in f2e.columns:
        if col_name not in col_map:
            new_col_idx = len(headers) + 1
            ws.cell(row=1, column=new_col_idx, value=col_name)
            col_map[col_name] = new_col_idx
            headers.append(col_name)

    existing_row_count = ws.max_row - 1  # excluding header

    # Update existing rows (preserve formatting, just change values)
    for df_idx in range(min(existing_row_count, len(f2e))):
        excel_row = df_idx + 2  # 1-based, skip header
        for col_name in f2e.columns:
            if col_name in col_map:
                val = f2e.iloc[df_idx][col_name]
                if pd.isna(val):
                    val = None
                ws.cell(row=excel_row, column=col_map[col_name], value=val)

    # Append new rows (beyond original row count)
    for df_idx in range(existing_row_count, len(f2e)):
        excel_row = df_idx + 2
        for col_name in f2e.columns:
            if col_name in col_map:
                val = f2e.iloc[df_idx][col_name]
                if pd.isna(val):
                    val = None
                ws.cell(row=excel_row, column=col_map[col_name], value=val)

    wb.save(output_path)
    print(f"\nSaved merged database to: {output_path}")
    print(f"Total projects: {len(f2e)}")


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Merge fresh CATF data into F2E CCS database')
    parser.add_argument('catf_file', help='Path to fresh CATF download (.xlsx)')
    parser.add_argument('f2e_file', help='Path to existing F2E database (.xlsx)')
    parser.add_argument('--output', '-o', default=None, help='Output path (default: overwrites f2e_file)')
    args = parser.parse_args()

    output = args.output or args.f2e_file
    merge(args.catf_file, args.f2e_file, output)
